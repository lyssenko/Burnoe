import re
import pandas as pd
import csv
import io
import os
import logging
from dotenv import load_dotenv
from sqlite3 import IntegrityError
from db_session import SessionLocal
from init_db import Sensor, Measurement, User
from datetime import datetime, timedelta
from flask import Flask, request, render_template, redirect, url_for
from sqlalchemy.dialects.sqlite import insert as sqlite_insert
from flask import session, flash, make_response
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from functools import wraps
from sensor_labels import SENSOR_LABELS, UNIT_LABELS
from comparison_utils import (
    get_common_time_series,
    get_measurements,
    get_sensor_names,
    handle_uploaded_file,
    parse_date_range,
    group_measurements,
    save_virtual_averages,
)


load_dotenv()
app = Flask(__name__)
secret = os.getenv("SECRET_KEY")
if not secret:
    raise RuntimeError("SECRET_KEY не найден в переменных окружения!")
app.secret_key = secret
app.permanent_session_lifetime = timedelta(minutes=10)
project_root = os.path.dirname(os.path.abspath(__file__))
log_path = os.path.join(project_root, "deletion.log")
logger = logging.getLogger("deletion_log")
handler = logging.FileHandler(log_path, encoding="utf-8")
formatter = logging.Formatter("%(asctime)s - %(message)s")
handler.setFormatter(formatter)
logger.setLevel(logging.INFO)
logger.addHandler(handler)


def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        session_cookie_name = app.config.get("SESSION_COOKIE_NAME", "session")
        if "username" not in session:
            if request.cookies.get(session_cookie_name):
                flash("Ваша сессия истекла, войдите снова", "warning")
            else:
                flash("Необходимо войти в систему", "warning")
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated_function


@app.context_processor
def inject_user():
    def sensor_label(name):
        return SENSOR_LABELS.get(name, name)

    def unit_label(unit):
        return UNIT_LABELS.get(unit, unit)

    return dict(
        username=session.get("username"),
        sensor_label=sensor_label,
        unit_label=unit_label,
    )


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]
        with SessionLocal() as db:
            user = db.query(User).filter_by(username=username).first()
            if user and user.check_password(password):
                session["username"] = username
                session.permanent = True
                flash("Вы вошли в систему", "success")
                return redirect(url_for("index"))
            else:
                flash("Неверный логин или пароль", "error")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.pop("username", None)
    flash("Вы вышли из системы", "info")
    return redirect(url_for("index"))


@app.route("/", methods=["GET"])
def index():
    return render_template("index.html")


@app.route("/upload", methods=["POST"])
@login_required
def upload():
    files = request.files.getlist("dataFile")
    if not files:
        return "Файл(ы) не выбраны", 400

    errors = []
    total_inserted = 0
    date = '2000-01-01'
    with SessionLocal() as db:
        for file in files:
            if file and file.filename:
                try:
                    inserted, date = handle_uploaded_file(file, db, errors)
                    total_inserted += inserted
                except Exception as e:
                    errors.append(f"Ошибка обработки {file.filename}: {e}")

        if errors:
            db.rollback()
            msg = (
                f"Загружено {total_inserted} измерений, но возникли ошибки:\n"
                + "\n".join(errors[:10])
            )
            if len(errors) > 10:
                msg += f"\n... и ещё {len(errors) - 10} ошибок скрыто."
            return msg, 400
    
        try:
            start_dt, end_dt = parse_date_range(date, date)
            save_virtual_averages(db, start_dt, end_dt)
        except Exception as e:
            db.rollback()
            return f"Ошибка при расчёте средних значений: {e}", 500

        db.commit()
    return f"Данные успешно загружены. Всего записей: {total_inserted}."


@app.route("/upload_forecast", methods=["POST"])
@login_required
def upload_forecast():
    files = request.files.getlist("forecastFile")
    if not files:
        return "Файлы прогноза не выбраны", 400

    total_inserted = 0
    with SessionLocal() as db:
        for file in files:
            if not file or not file.filename:
                continue

            filename_lower = file.filename.lower()
            if "energy" in filename_lower:
                sensor_name = "Forecast Energy"
                sensor_type = "energy_active"
                unit = "kWh"
            elif "irrad" in filename_lower:
                sensor_name = "Forecast Radiation"
                sensor_type = "radiation"
                unit = "W/m2"
            else:
                print(
                    f"Пропущен файл: {file.filename} — не содержит 'energy' или 'irrad'"
                )
                continue

            try:
                df = pd.read_csv(file)
                df.columns = [col.strip().lower() for col in df.columns]
                if "time" in df.columns and "p" in df.columns:
                    df = df.rename(columns={"p": "radiation"})
                elif "time" in df.columns and "rad" in df.columns:
                    df = df.rename(columns={"rad": "radiation"})
                elif "time" in df.columns and "radiation" in df.columns:
                    pass
                else:
                    print(f"Ошибка структуры в файле {file.filename}")
                    continue
            except Exception as e:
                print(f"Ошибка чтения {file.filename}: {e}")
                continue

            try:
                base_name = file.filename.split("/")[-1]
                date_str = re.search(r"(\d{4}-\d{2}-\d{2})", base_name)
                if date_str:
                    forecast_date = datetime.strptime(
                        date_str.group(1), "%Y-%m-%d"
                    ).date()
                else:
                    forecast_date = datetime.today().date()
            except Exception:
                forecast_date = datetime.today().date()

            sensor = db.query(Sensor).filter_by(sensor_name=sensor_name).first()
            if sensor:
                sensor_id = sensor.sensor_id
            else:
                sensor = Sensor(
                    sensor_name=sensor_name, sensor_type=sensor_type, unit=unit
                )
                db.add(sensor)
                db.flush()
                sensor_id = sensor.sensor_id

            for _, row in df.iterrows():
                try:
                    time_obj = datetime.strptime(row["time"], "%H:%M:%S").time()
                    dt = datetime.combine(forecast_date, time_obj)
                    value = float(row["radiation"])

                    stmt = (
                        sqlite_insert(Measurement)
                        .values(sensor_id=sensor_id, measurement_time=dt, value=value)
                        .on_conflict_do_update(
                            index_elements=["sensor_id", "measurement_time"],
                            set_={"value": value},
                        )
                    )
                    db.execute(stmt)
                    total_inserted += 1
                except Exception as e:
                    print(f"Ошибка прогноза в строке: {e}")

        try:
            db.commit()
        except IntegrityError as e:
            db.rollback()
            return f"Ошибка при коммите прогноза: {e}", 500

    return f"Загружено: {total_inserted} прогнозных значений."


@app.route("/sensors")
@login_required
def show_sensors():
    with SessionLocal() as db:
        sensors = (
            db.query(Sensor).filter_by(visible=True).order_by(Sensor.sensor_id).all()
        )
    return render_template("sensors.html", sensors=sensors)


@app.route("/data", methods=["GET", "POST"])
def show_data():
    selected_sensor_id = request.args.get("sensor_id")
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")
    interval = int(request.args.get("interval", 15))

    if not start_date and not end_date and not selected_sensor_id:
        default_start = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
        default_end = (datetime.now()).strftime("%Y-%m-%d")
        return redirect(
            url_for("show_data", start_date=default_start, end_date=default_end)
        )

    with SessionLocal() as db:
        sensors = db.query(Sensor).filter_by(visible=True).all()
        query = db.query(Measurement)
        if selected_sensor_id and selected_sensor_id.isdigit():
            query = query.filter(Measurement.sensor_id == int(selected_sensor_id))

        start_dt, end_dt = parse_date_range(start_date, end_date)
        query = query.filter(Measurement.measurement_time >= start_dt)
        query = query.filter(Measurement.measurement_time <= end_dt)
        measurements = query.order_by(Measurement.measurement_time).all()

        grouped = group_measurements(measurements, interval)
        chart_labels = [dt.strftime("%Y-%m-%d %H:%M") for dt in sorted(grouped)]
        chart_values = [grouped[dt] for dt in sorted(grouped)]

    return render_template(
        "data.html",
        sensors=sensors,
        measurements=measurements,
        selected_sensor_id=selected_sensor_id,
        start_date=start_date,
        end_date=end_date,
        chart_labels=chart_labels,
        chart_values=chart_values,
        interval=interval,
    )


@app.route("/compare_select", methods=["GET"])
def compare_select():
    data_type = request.args.get("data_type", "radiation")

    from datetime import date, timedelta

    default_start = (date.today() - timedelta(days=1)).strftime("%Y-%m-%d")
    default_end = date.today().strftime("%Y-%m-%d")

    with SessionLocal() as db:
        actual_type_filter = (
            [data_type, "virtual"] if data_type == "radiation" else [data_type]
        )

        sensors_actual = (
            db.query(Sensor)
            .filter(
                ~Sensor.sensor_name.ilike("%forecast%"),
                Sensor.visible == True,
                Sensor.sensor_type.in_(actual_type_filter)
            )
            .order_by(Sensor.sensor_name)
            .all()
        )

        sensors_forecast = (
            db.query(Sensor)
            .filter(
                Sensor.sensor_name.ilike("%forecast%"),
                Sensor.visible == True,
                Sensor.sensor_type == data_type
            )
            .order_by(Sensor.sensor_name)
            .all()
        )

    return render_template(
        "compare_select.html",
        sensors_actual=sensors_actual,
        sensors_forecast=sensors_forecast,
        data_type=data_type,
        default_start=default_start,
        default_end=default_end,
    )


@app.route("/compare", methods=["GET"])
def compare():
    with SessionLocal() as db:
        try:
            actual_id = int(request.args.get("sensor_actual_id"))
            forecast_id = int(request.args.get("sensor_forecast_id"))
            start_date = request.args.get("start_date")
            end_date = request.args.get("end_date")
            interval = int(request.args.get("interval", 15))

            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            end_dt = (
                datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
                if end_date
                else start_dt + timedelta(days=1)
            )

            actual_name, forecast_name, actual_unit, forecast_unit = get_sensor_names(db, actual_id, forecast_id)

            actual_data = get_measurements(db, actual_id, start_dt, end_dt)
            forecast_data = get_measurements(db, forecast_id, start_dt, end_dt)

            grouped_actual = group_measurements(actual_data, interval)
            grouped_forecast = group_measurements(forecast_data, interval)

            all_times = sorted(
                set(grouped_actual.keys()).union(grouped_forecast.keys())
            )
            actual_values = [grouped_actual.get(dt, None) for dt in all_times]
            forecast_values = [grouped_forecast.get(dt, None) for dt in all_times]

            return render_template(
                "compare.html",
                chart_labels=[dt.strftime("%Y-%m-%d %H:%M") for dt in all_times],
                actual_values=actual_values,
                forecast_values=forecast_values,
                actual_name=actual_name,
                forecast_name=forecast_name,
                interval=interval,
                actual_unit=actual_unit,
                forecast_unit=forecast_unit,
            )
        except Exception as e:
            print(f"Ошибка в /compare: {e}")
            return redirect(url_for("compare_select"))


@app.route("/compare_table", methods=["GET"])
def compare_table():
    from collections import defaultdict

    with SessionLocal() as db:
        sensor_actual_id = request.args.get("sensor_actual_id")
        sensor_forecast_id = request.args.get("sensor_forecast_id")
        start_date = request.args.get("start_date")
        end_date = request.args.get("end_date")
        interval = int(request.args.get("interval", 15))

        if not (sensor_actual_id and sensor_forecast_id and start_date):
            return redirect(url_for("compare_select"))

        try:
            sensor_actual_id = int(sensor_actual_id)
            sensor_forecast_id = int(sensor_forecast_id)
        except ValueError:
            return "Некорректный формат ID сенсора", 400

        try:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
        except ValueError:
            return "Неверный формат начальной даты", 400

        if end_date:
            try:
                end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
            except ValueError:
                return "Неверный формат конечной даты", 400
        else:
            end_dt = start_dt + timedelta(days=1)

        actual_name = (
            "Среднее"
            if sensor_actual_id == -1
            else db.get(Sensor, sensor_actual_id).sensor_name
        )
        
        from comparison_utils import get_sensor_names
        actual_name, forecast_name, actual_unit, forecast_unit = get_sensor_names(
            db, sensor_actual_id, sensor_forecast_id
        )

        if sensor_actual_id == -1:
            from comparison_utils import get_avg_measurements_for_all

            actual_data = get_avg_measurements_for_all(db, start_dt, end_dt)
        else:
            actual_data = (
                db.query(Measurement)
                .filter(
                    Measurement.sensor_id == sensor_actual_id,
                    Measurement.measurement_time >= start_dt,
                    Measurement.measurement_time < end_dt,
                )
                .order_by(Measurement.measurement_time)
                .all()
            )

        forecast_data = (
            db.query(Measurement)
            .filter(
                Measurement.sensor_id == sensor_forecast_id,
                Measurement.measurement_time >= start_dt,
                Measurement.measurement_time < end_dt,
            )
            .order_by(Measurement.measurement_time)
            .all()
        )

    def bucket_time(dt, interval):
        minute_bucket = (dt.minute // interval) * interval
        return dt.replace(minute=minute_bucket, second=0, microsecond=0)

    actual_buckets = defaultdict(list)
    forecast_buckets = defaultdict(list)
    for m in actual_data:
        b_time = bucket_time(m.measurement_time, interval)
        actual_buckets[b_time].append(m.value)
    for m in forecast_data:
        b_time = bucket_time(m.measurement_time, interval)
        forecast_buckets[b_time].append(m.value)

    all_buckets = sorted(set(actual_buckets.keys()) | set(forecast_buckets.keys()))

    grouped_rows = defaultdict(list)
    daily_totals = []

    for b_time in all_buckets:
        a_vals = actual_buckets.get(b_time, [])
        f_vals = forecast_buckets.get(b_time, [])

        a_avg = round(sum(a_vals) / len(a_vals), 3) if a_vals else ""
        f_avg = round(sum(f_vals) / len(f_vals), 3) if f_vals else ""

        a_avg = a_avg if isinstance(a_avg, (int, float)) and a_avg > 0 else 0
        f_avg = f_avg if isinstance(f_avg, (int, float)) and f_avg > 0 else 0
        err = (a_avg - f_avg) if (a_avg != "" and f_avg != "") else ""
        percent = (
            ((err / a_avg) * 100)
            if (a_avg != "" and f_avg != "" and a_avg != 0)
            else ""
        )
        date_key = b_time.strftime("%Y-%m-%d")
        row = {
            "date": date_key,
            "time": b_time.strftime("%Y-%m-%d %H:%M"),
            "actual": a_avg,
            "forecast": f_avg,
            "error": round(err, 3) if err != "" else "",
            "percent": round(percent, 2) if percent != "" else "",
        }
        grouped_rows[date_key].append(row)

    daily_totals_dict = {}
    for date, rows in grouped_rows.items():
        actual_sum = sum(
            float(row["actual"]) for row in rows if row["actual"] not in ("", None)
        )
        forecast_sum = sum(
            float(row["forecast"]) for row in rows if row["forecast"] not in ("", None)
        )
        abs_error = actual_sum - forecast_sum
        percent_error = (abs_error / actual_sum * 100) if actual_sum else 0.0

        daily_totals_dict[date] = {
            "date": date,
            "actual": round(actual_sum, 3),
            "forecast": round(forecast_sum, 3),
            "abs_error": round(abs_error, 3),
            "percent_error": round(percent_error, 2)
        }

    daily_totals = list(daily_totals_dict.values())


    total_actual_sum = sum(
        row["actual"]
        for rows in grouped_rows.values()
        for row in rows
        if isinstance(row["actual"], (int, float))
    )
    total_forecast_sum = sum(
        row["forecast"]
        for rows in grouped_rows.values()
        for row in rows
        if isinstance(row["forecast"], (int, float))
    )

    abs_error = total_actual_sum - total_forecast_sum
    percent_error = (abs_error / total_actual_sum * 100) if total_actual_sum else None

    return render_template(
        "compare_table.html",
        grouped_rows=grouped_rows,
        daily_totals=daily_totals,
        actual_name=actual_name,
        forecast_name=forecast_name,
        total_actual_sum=round(total_actual_sum, 3),
        total_forecast_sum=round(total_forecast_sum, 3),
        abs_error=round(abs_error, 3),
        percent_error=round(percent_error, 2) if percent_error is not None else "",
        interval=interval,
        actual_unit=actual_unit,
        forecast_unit=forecast_unit,
    )


@app.route("/compare_table/export")
def export_compare_table():
    db = SessionLocal()
    actual_id = int(request.args.get("sensor_actual_id"))
    forecast_id = int(request.args.get("sensor_forecast_id"))
    start_date = datetime.strptime(request.args.get("start_date"), "%Y-%m-%d")
    end_date = datetime.strptime(request.args.get("end_date"), "%Y-%m-%d") + timedelta(
        days=1
    )

    actual_data = get_measurements(db, actual_id, start_date, end_date)
    forecast_data = get_measurements(db, forecast_id, start_date, end_date)
    labels, actual_dict, forecast_dict = get_common_time_series(
        actual_data, forecast_data
    )

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(["Время", "Факт", "Прогноз", "Ошибка", "Ошибка (%)"])

    for t in labels:
        a = actual_dict.get(t)
        f = forecast_dict.get(t)
        diff = (a - f) if a is not None and f is not None else ""
        percent = (abs(diff) / a * 100) if a and f else ""
        writer.writerow([t.strftime("%Y-%m-%d %H:%M:%S"), a, f, diff, percent])

    response = make_response(buffer.getvalue())
    response.headers["Content-Disposition"] = "attachment; filename=comparison.csv"
    response.headers["Content-type"] = "text/csv"
    return response


@app.route("/compare_table/export_excel")
def export_compare_table_excel():
    db = SessionLocal()
    actual_id = int(request.args.get("sensor_actual_id"))
    forecast_id = int(request.args.get("sensor_forecast_id"))
    start_date = datetime.strptime(request.args.get("start_date"), "%Y-%m-%d")
    end_date = datetime.strptime(request.args.get("end_date"), "%Y-%m-%d") + timedelta(
        days=1
    )

    actual_data = get_measurements(db, actual_id, start_date, end_date)
    forecast_data = get_measurements(db, forecast_id, start_date, end_date)
    labels, actual_dict, forecast_dict = get_common_time_series(
        actual_data, forecast_data
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Сравнение"

    ws.append(["Время", "Факт", "Прогноз", "Ошибка", "Ошибка (%)"])

    for t in labels:
        a = actual_dict.get(t)
        f = forecast_dict.get(t)
        diff = (a - f) if a is not None and f is not None else ""
        percent = (abs(diff) / a * 100) if a and f else ""
        ws.append(
            [
                t.strftime("%Y-%m-%d %H:%M:%S"),
                a if a is not None else "",
                f if f is not None else "",
                diff if diff != "" else "",
                round(percent, 2) if percent != "" else "",
            ]
        )

    for col in ws.columns:
        max_len = max(
            len(str(cell.value)) if cell.value is not None else 0 for cell in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = make_response(buffer.read())
    response.headers["Content-Disposition"] = "attachment; filename=comparison.xlsx"
    response.headers["Content-type"] = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    return response


@app.route("/admin/sensors", methods=["GET", "POST"])
@login_required
def admin_sensors():
    with SessionLocal() as db:
        if request.method == "POST":
            visible_ids = set(map(int, request.form.getlist("visible_sensor")))
            all_sensors = db.query(Sensor).all()
            for sensor in all_sensors:
                sensor.visible = sensor.sensor_id in visible_ids
            db.commit()
            flash("Настройки видимости сохранены.", "success")
            return redirect(url_for("admin_sensors"))

        sensors = db.query(Sensor).order_by(Sensor.sensor_id).all()
    return render_template("admin_sensors.html", sensors=sensors)


@app.route("/delete_measurements", methods=["POST"])
def delete_measurements():
    if "username" not in session:
        flash("Необходима авторизация", "danger")
        return redirect(url_for("login"))

    username = session.get("username")
    sensor_id = request.form.get("sensor_id")
    start_date = request.form.get("start_date")
    end_date = request.form.get("end_date")

    if not sensor_id or not start_date:
        flash("Неверные параметры удаления", "danger")
        return redirect(url_for("show_data"))

    try:
        sensor_id = int(sensor_id)
    except ValueError:
        flash("Некорректный ID сенсора", "danger")
        return redirect(url_for("show_data"))

    start_dt, end_dt = parse_date_range(start_date, end_date)

    with SessionLocal() as db:
        query = db.query(Measurement).filter(Measurement.sensor_id == sensor_id)

        if start_dt:
            query = query.filter(Measurement.measurement_time >= start_dt)
        if end_dt:
            query = query.filter(Measurement.measurement_time <= end_dt)

        deleted_rows = query.delete(synchronize_session=False)
        db.commit()

    logger.info(f"Пользователь '{username}' удалил {deleted_rows} записей (sensor_id={sensor_id}, от {start_date} до {end_date})")

    flash(f"Удалено {deleted_rows} измерений.", "success")
    return redirect(url_for("show_data", sensor_id=sensor_id, start_date=start_date, end_date=end_date))


if __name__ == "__main__":
    app.run(debug=True, use_reloader=False)
